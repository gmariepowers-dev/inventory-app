from flask import Flask, request, render_template, redirect, url_for, abort, flash, jsonify, Response
from models import db, Item, User, AuditLog

import os
import time
import shutil

from io import BytesIO
from functools import wraps

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

from barcode import Code128
from barcode.writer import ImageWriter 

from flask_login import (
    LoginManager,
    login_user,
    logout_user,
    login_required,
    current_user,
)

from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename

# --------------------
# Config
# --------------------
ALLOWED_EXTENSIONS = {"png", "jpg", "jpeg", "webp"}

def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS

# --------------------
# App setup
# --------------------
app = Flask(__name__, static_folder="static", template_folder="templates", instance_relative_config=True)

app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "dev-secret-key")
app.config["MAX_CONTENT_LENGTH"] = 5 * 1024 * 1024 #5MB upload limit

os.makedirs(app.instance_path, exist_ok=True)
db_path = os.path.join(app.instance_path, "inventory.db")

database_url = os.getenv("DATABASE_URL")

if database_url:
    database_url = database_url.replace("postgres://", "postgresql://", 1)
    app.config["SQLALCHEMY_DATABASE_URI"] = database_url
else:
    app.config["SQLALCHEMY_DATABASE_URI"] = f"sqlite:///{db_path}"

app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

db.init_app(app)

# --------------------
# Login manager
# --------------------
login_manager = LoginManager(app)
login_manager.login_view = "login"

@login_manager.user_loader
def load_user(user_id):
    return db.session.get(User, int(user_id))

# --------------------
# Admin guard
# --------------------
def admin_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.is_admin:
            abort(403)
        return view(*args, **kwargs)
    return wrapped

# --------------------
# Audit logger
# --------------------
def log_audit(action, target=None, details=None):
    db.session.add(
        AuditLog(
            actor_id=current_user.id if current_user.is_authenticated else None,
            action=action,
            target=target,
            details=details
        )
    )
    db.session.commit()

# --------------------
# Auth
# --------------------
@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        user = User.query.filter_by(username=request.form["username"]).first()

        if user and check_password_hash(user.password_hash, request.form["password"]):
            login_user(user)
            log_audit("login", target=user.username)
            return redirect(request.args.get("next") or url_for("index"))
        
        flash("Invalid username or password", "error")

    return render_template("login.html")

@app.route("/logout")
@login_required
def logout():
    log_audit("logout", target=current_user.username)
    logout_user()
    return redirect(url_for("login"))

# --------------------
# Inventory
# --------------------
@app.route("/", methods=["GET", "POST"])
@login_required
def index():

    if request.method == "POST":
        sku = request.form["sku"]

        # Validation
        if not sku:
            flash("SKU is required", "error")
            return redirect(url_for("index"))

        if not request.form["name"]:
            flash("Item name is required", "error")
            return redirect(url_for("index"))
        
        if len(request.form["name"]) > 200:
            flash("Name too long", "error")
            return redirect(url_for("index"))
        
        if len(sku) > 100:
            flash("SKU too long", "error")
            return redirect(url_for("index"))

        if Item.query.filter_by(sku=sku).first():
            flash("SKU already exists", "error")
            return redirect(url_for("index"))
        
        # Quantity validation
        qty = int(request.form.get("quantity") or 0)
        threshold = int(request.form.get("low_stock_threshold") or 5)

        if qty < 0:
            flash("Quantity cannot be negative", "error")
            return redirect(url_for("index"))
        
        

       # Barcode
        barcode_dir = os.path.join(app.static_folder, "barcodes")
        os.makedirs(barcode_dir, exist_ok=True)

        barcode = Code128(sku, writer=ImageWriter())
        barcode.save(
            os.path.join(barcode_dir, sku),
            options={
                "module_width": 0.28,
                "module_height": 28,
                "quiet_zone": 7,
                "font_size": 12,
                "text_distance": 6,
                "dpi": 300,
                "write_text": True,
            }
        )

        barcode_path = f"/static/barcodes/{sku}.png"

        # Image upload
        image = request.files.get("image")
        image_path = None

        if image and image.filename:
            if not allowed_file(image.filename):
                flash("Invalid image type", "error")
                return redirect(url_for("index"))
            
            image_dir = os.path.join(app.static_folder, "uploads")
            os.makedirs(image_dir, exist_ok=True)

            filename = secure_filename(f"{sku}_{image.filename}")
            image.save(os.path.join(image_dir, filename))
            image_path = f"/static/uploads/{filename}"

        # Create item
        item = Item(
            name=request.form["name"],
            sku=sku,
            quantity=qty,
            low_stock_threshold=threshold,  
            dimensions=request.form.get("dimensions"),
            weight=request.form.get("weight"),
            colorways=request.form.get("colorways"),
            manufacturer=request.form.get("manufacturer"),
            cost_price=float(request.form.get("cost_price") or 0),
            retail_price=float(request.form.get("retail_price") or 0),
            barcode_path=barcode_path,
            barcode_value=sku,
            image_path=image_path
        )

        db.session.add(item)
        db.session.commit()

        log_audit("create_item", target=item.sku, details=item.name)

        flash("Item added successfully", "success")
        return redirect(url_for("index"))

    items = Item.query.all()
    return render_template("index.html", items=items)

# --------------------
# Item APIs
# --------------------
@app.route("/item/<int:item_id>")
@login_required
def get_item(item_id):
    return jsonify(Item.query.get_or_404(item_id).to_dict())

@app.route("/edit-item/<int:item_id>", methods=["POST"])
@login_required
def edit_item(item_id):
    item = Item.query.get_or_404(item_id)
    data = request.get_json()

    item.manufacturer = data.get("manufacturer")
    item.dimensions = data.get("dimensions")
    item.colorways = data.get("colorways")
    item.cost_price = float(data.get("cost_price") or 0)
    item.retail_price = float(data.get("retail_price") or 0)
    item.weight = data.get("weight")

    db.session.commit()
    log_audit("edit_item", target=item.sku)

    return jsonify({"success": True})

@app.route("/set-quantity/<int:item_id>", methods=["POST"])
@login_required
def set_quantity(item_id):
    item = Item.query.get_or_404(item_id)
    data = request.get_json()

    new_qty = int(data.get("quantity", 0))
    if new_qty < 0:
        return jsonify({"error": "Invalid quantity"}), 400

    item.quantity = new_qty
    db.session.commit()

    log_audit("set_quantity", target=item.sku, details=str(new_qty))

    return jsonify({"success": True})

@app.route("/adjust-quantity/<int:item_id>", methods=["POST"])
@login_required
def adjust_quantity(item_id):
    item = Item.query.get_or_404(item_id)
    data = request.get_json()

    change = int(data.get("change", 0))
    item.quantity = max(0, (item.quantity or 0) + change)

    db.session.commit()

    log_audit("adjust_quantity", target=item.sku, details=f"change {change}, new qty {item.quantity}")

    return jsonify({"new_quantity": item.quantity})

@app.route("/delete-item/<int:item_id>", methods=["POST"])
@login_required
def delete_item(item_id):
    item = Item.query.get_or_404(item_id)

    db.session.delete(item)
    db.session.commit()

    log_audit("delete_item", target=item.sku)

    flash("Item deleted", "success")
    return redirect(url_for("index"))

# --------------------
# Summary
# --------------------
@app.route("/summary")
@login_required
def summary():
    items = Item.query.all()

    low_stock_threshold = 5

    total_items = len(items)
    total_retail_value = sum((item.quantity or 0) * (item.retail_price or 0) for item in items)
    total_cost_value = sum((item.quantity or 0) * (item.cost_price or 0) for item in items)

    labels = [item.name for item in items]
    quantities = [item.quantity or 0 for item in items]
    values = [(item.quantity or 0) * (item.cost_price or 0) for item in items]

    item_values = []
    for item in items:
        qty = item.quantity or 0
        cost_price = item.cost_price or 0
        retail_price = item.retail_price or 0

        item_values.append({
            "name": item.name,
            "sku": item.sku,
            "quantity": qty,
            "cost_price": cost_price,
            "retail_price": retail_price,
            "value": qty * cost_price,
            "is_low_stock": qty <= (item.low_stock_threshold or 0)
        })

    low_stock_items = [
    item for item in items
    if (item.quantity or 0) <= (item.low_stock_threshold or 0)
]

    return render_template(
        "summary.html",
        total_items=total_items,
        total_retail_value=total_retail_value,
        total_cost_value=total_cost_value,
        labels=labels,
        quantities=quantities,
        values=values,
        item_values=item_values,
        low_stock_items=low_stock_items,
        low_stock_threshold=low_stock_threshold
    )
@app.route("/summary/export")
@login_required
def export_inventory():
    items = Item.query.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory Summary"

    # Styles
    header_fill = PatternFill("solid", fgColor="2B51A3")
    header_font = Font(color="FFFFFF", bold=True)
    subheader_fill = PatternFill("solid", fgColor="EAF0FB")
    bold_font = Font(bold=True)
    low_stock_fill = PatternFill("solid", fgColor="FEE2E2")
    thin_gray = Side(style="thin", color="D1D5DB")

    # Summary stats
    total_items = len(items)
    total_quantity = sum(item.quantity or 0 for item in items)
    total_retail_value = sum((item.quantity or 0) * (item.retail_price or 0) for item in items)
    total_cost_value = sum((item.quantity or 0) * (item.cost_price or 0) for item in items)

    ws["A1"] = "Inventory Export"
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    ws["A1"].fill = header_fill

    ws.merge_cells("A1:H1")

    ws["A3"] = "Total Items"
    ws["B3"] = total_items

    ws["A4"] = "Total Quantity"
    ws["B4"] = total_quantity

    ws["A5"] = "Total Retail Value"
    ws["B5"] = total_retail_value

    ws["A6"] = "Our Price Value"
    ws["B6"] = total_cost_value

    for cell in ["A3", "A4", "A5", "A6"]:
        ws[cell].font = bold_font

    ws["B5"].number_format = '$#,##0.00'
    ws["B6"].number_format = '$#,##0.00'

    # Table header
    start_row = 8
    headers = [
        "Item Name",
        "SKU",
        "Quantity",
        "Low Stock Threshold",
        "Cost Price",
        "Retail Price",
        "Total Cost Value",
        "Total Retail Value",
        "Low Stock"
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=thin_gray)

    # Data rows
    for idx, item in enumerate(items, start=start_row + 1):
        qty = item.quantity or 0
        cost_price = item.cost_price or 0
        retail_price = item.retail_price or 0
        threshold = item.low_stock_threshold or 5
        is_low_stock = qty <= threshold

        row_values = [
            item.name,
            item.sku,
            qty,
            threshold,
            cost_price,
            retail_price,
            qty * cost_price,
            qty * retail_price,
            "YES" if is_low_stock else "NO"
        ]

        for col_num, value in enumerate(row_values, 1):
            cell = ws.cell(row=idx, column=col_num, value=value)
            cell.alignment = Alignment(vertical="center")

            if col_num in [5, 6, 7, 8]:
                cell.number_format = '$#,##0.00'

            if is_low_stock:
                cell.fill = low_stock_fill

    # Column widths
    widths = {
        "A": 24,
        "B": 18,
        "C": 12,
        "D": 20,
        "E": 14,
        "F": 14,
        "G": 18,
        "H": 18,
        "I": 12,
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A9"

    # Export
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=inventory_export.xlsx"
        }
    )

# --------------------
# Labels
# --------------------
@app.route("/labels")
@login_required
def labels_page():
    return redirect(url_for("labels_builder"))


@app.route("/labels/builder", methods=["GET", "POST"])
@login_required
def labels_builder():
    items = Item.query.order_by(Item.name.asc()).all()

    if request.method == "POST":
        selected_ids = request.form.getlist("selected_items")

        if not selected_ids:
            flash("Select at least one item.", "error")
            return redirect(url_for("labels_builder"))

        selected_items = (
            Item.query
            .filter(Item.id.in_(selected_ids))
            .order_by(Item.name.asc())
            .all()
        )

        return render_template("labels_print.html", items=selected_items)

    return render_template("labels_builder.html", items=items)


@app.route("/labels/item/<int:item_id>")
@login_required
def label_single_item(item_id):
    item = Item.query.get_or_404(item_id)
    return render_template("labels_print.html", items=[item])

# --------------------
# Admin
# --------------------
@app.route("/admin")
@login_required
@admin_required
def admin():
    users = User.query.all()
    logs = AuditLog.query.order_by(AuditLog.id.desc()).limit(10).all()

    return render_template(
        "admin.html",
        users=users,
        logs=logs
    )

@app.route("/admin/add-user", methods=["GET", "POST"])
@login_required
@admin_required
def add_user():
    if request.method == "POST":
        username = request.form["username"]
        password = request.form["password"]
        role = request.form["role"]

        if User.query.filter_by(username=username).first():
            flash("User already exists", "error")
            return redirect(url_for("admin"))

        new_user = User(
            username=username,
            password_hash=generate_password_hash(password),
            role=role
        )

        db.session.add(new_user)
        db.session.commit()

        log_audit("create_user", target=username)
        flash("User created", "success")

        return redirect(url_for("admin"))

    return render_template("add_user.html")

@app.route("/admin/users/edit/<int:user_id>", methods=["POST"])
@login_required
@admin_required
def edit_user(user_id):

    user = User.query.get_or_404(user_id)

    new_role = request.form["role"]
    user.role = new_role

    db.session.commit()

    log_audit("edit_user", target=user.username, details=f"role -> {new_role}")

    flash("User updated", "success")
    return redirect(url_for("admin"))


@app.route("/admin/users/delete/<int:user_id>", methods=["POST"])
@login_required
@admin_required
def delete_user(user_id):

    user = User.query.get_or_404(user_id)


    if user.username == "admin":
        flash("Cannot delete main admin", "error")
        return redirect(url_for("admin"))

    db.session.delete(user)
    db.session.commit()

    log_audit("delete_user", target=user.username)

    flash("User deleted", "success")
    return redirect(url_for("admin"))

# --------------------
# Audit Logs Page
# --------------------
@app.route("/admin/logs")
@login_required
@admin_required
def view_audit_logs():
    logs = AuditLog.query.order_by(AuditLog.id.desc()).limit(200).all()
    return render_template("audit_logs.html", logs=logs)

@app.route("/admin/audit/export")
@login_required
@admin_required
def export_audit_logs():
    logs = AuditLog.query.order_by(AuditLog.created_at.desc()).all()

    def generate():
        yield "timestamp,user,action,item\n"
        for log in logs:
            yield f"{log.created_at},{log.actor_id},{log.action},{log.target}\n"

    return Response(generate(), mimetype="text/csv",
                    headers={"Content-Disposition": "attachment;filename=audit_logs.csv"})

@app.route("/admin/users/reset-password/<int:user_id>", methods=["POST"])
@login_required
@admin_required
def reset_user_password(user_id):
    user = User.query.get_or_404(user_id)
    new_password = request.form.get("new_password", "").strip()

    if len(new_password) < 8:
        flash("Password must be at least 8 characters", "error")
        return redirect(url_for("admin"))

    user.password_hash = generate_password_hash(new_password)
    db.session.commit()

    log_audit("reset_password", target=user.username)
    flash(f"Password reset for {user.username}", "success")
    return redirect(url_for("admin"))

# --------------------
# Backup
# --------------------
@app.route("/admin/backup")
@login_required
@admin_required
def backup_db():
    backup_path = os.path.join(app.instance_path, f"backup_{int(time.time())}.db")
    shutil.copy(db_path, backup_path)
    flash("Backup created", "success")
    return redirect(url_for("admin"))


# --------------------
# Barcode Scanner
# --------------------
@app.route("/scan/<barcode>")
@login_required
def scan_item(barcode):
    barcode = barcode.strip()
    print("SCANNED BARCODE:", repr(barcode))

    item = Item.query.filter(
        (Item.sku == barcode) | (Item.barcode_value == barcode)
    ).first()

    print("FOUND ITEM:", item)

    if not item:
        return jsonify({"found": False})

    return jsonify({
        "found": True,
        "id": item.id
    })

# --------------------
# Create default admin
# --------------------
def create_admin_if_missing():
    if not User.query.filter_by(username="admin").first():
        default_password = os.environ.get("DEFAULT_ADMIN_PASSWORD", "admin123")

        admin = User(
            username="admin",
            password_hash=generate_password_hash(default_password),
            role="admin"
        )

        db.session.add(admin)
        db.session.commit()

# --------------------
# Run
# --------------------
if __name__ == "__main__":
    with app.app_context():
        db.create_all()
        create_admin_if_missing()

    app.run(debug=True)